import openpyxl, qrcode, json, os, re, shutil, socket

# 以脚本所在目录为基准，全部使用相对路径，便于迁移到服务器或 GitHub
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(BASE_DIR, '人员.xlsx')
OUT = os.path.join(BASE_DIR, 'outputs')
QR_DIR = os.path.join(OUT, 'qrcodes')
CARD_DIR = os.path.join(OUT, 'cards')
CARDS_SRC = os.path.join(BASE_DIR, '企业微信名片')
API_BASE = 'https://mantledb.sh/v2/eugooo/submissions'
SERVER_PORT = 3000

# ========== 部署模式配置 ==========
# 'local'  = 局域网部署，QR 码使用本机 IP
# 'github' = GitHub Pages 部署，QR 码使用 GitHub Pages URL
DEPLOY_MODE = 'github'

# GitHub Pages 配置（部署到 GitHub 时修改这里）
# 格式: https://<用户名>.github.io/<仓库名>
GITHUB_PAGES_URL = 'https://hide-star.github.io/eugooo-share'

# 自动获取本机局域网 IP
def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return 'localhost'

if DEPLOY_MODE == 'github':
    BASE_URL = GITHUB_PAGES_URL
    print(f'部署模式: GitHub Pages')
else:
    LOCAL_IP = get_local_ip()
    BASE_URL = f'http://{LOCAL_IP}:{SERVER_PORT}'
    print(f'部署模式: 局域网')
    print(f'本机 IP: {LOCAL_IP}')
print(f'QR 码基础 URL: {BASE_URL}')

os.makedirs(QR_DIR, exist_ok=True)
os.makedirs(CARD_DIR, exist_ok=True)

available_cards = {}
if os.path.exists(CARDS_SRC):
    for f in os.listdir(CARDS_SRC):
        if f.lower().endswith(('.jpg', '.jpeg', '.png')):
            name = os.path.splitext(f)[0]
            src_path = os.path.join(CARDS_SRC, f)
            dst_path = os.path.join(CARD_DIR, f)
            shutil.copy2(src_path, dst_path)
            available_cards[name] = f'cards/{f}'
            print(f'  Card: {f}')

wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb['Sheet1']
form_url = ws.cell(row=1, column=4).value

people = []
for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    eid = ws.cell(row=row, column=2).value
    script = ws.cell(row=row, column=3).value
    if name and eid:
        people.append({
            'name': name.strip(),
            'id': str(eid).strip(),
            'script': (script or '').strip()
        })

FORM_TPL = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, user-scalable=no">
<title>__NAME__ · EUGOOO 入驻</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:"Microsoft YaHei","PingFang SC",sans-serif;background:linear-gradient(180deg,#e8f0fe 0%,#f5f7fa 40%);min-height:100vh;color:#222;padding-bottom:40px}
.top{background:linear-gradient(135deg,#1a73e8,#0d47a1);color:#fff;text-align:center;padding:32px 20px 24px;border-radius:0 0 24px 24px}
.top .label{font-size:13px;opacity:.85;margin-bottom:4px}
.top h2{font-size:22px;margin-bottom:6px}
.top .ref{font-size:14px;opacity:.9}
.top .ref strong{font-size:22px;display:block;margin-top:4px;letter-spacing:4px}
.card{background:#fff;margin:16px;border-radius:16px;padding:20px;box-shadow:0 2px 12px rgba(0,0,0,.06)}
.card h3{font-size:16px;color:#1a73e8;margin-bottom:10px}
.card .msg{white-space:pre-wrap;color:#555;font-size:13px;line-height:1.8;max-height:200px;overflow-y:auto;background:#f8faff;padding:12px;border-radius:8px}
.card-script{display:flex;gap:14px;align-items:flex-start}
.card-script>.wecom-col{flex:0 0 100px;text-align:center}
.card-script>.wecom-col img{width:100px;height:auto;border-radius:8px;display:block}
.card-script>.wecom-col .cname{font-size:12px;font-weight:700;color:#2e7d32;margin-top:4px}
.card-script>.wecom-col .cdesc{font-size:10px;color:#888;margin-top:1px}
.card-script>.script-col{flex:1;min-width:0}
@media (max-width:450px){.card-script{flex-direction:column}}
.field{margin-bottom:14px}
.field label{display:block;font-size:14px;color:#555;margin-bottom:6px}
.field label .req{color:#e53935}
.field input,.field select{width:100%;padding:12px;border:1px solid #ddd;border-radius:10px;font-size:15px;background:#fafafa;-webkit-appearance:none}
.field input:focus,.field select:focus{outline:none;border-color:#1a73e8;background:#fff}
.field .locked{background:#e8f0fe;color:#1a73e8;font-weight:700;border-color:#1a73e8}
.submit-btn{width:100%;background:linear-gradient(135deg,#1a73e8,#0d47a1);color:#fff;border:none;padding:16px;border-radius:12px;font-size:18px;font-weight:700;cursor:pointer;margin-top:8px;-webkit-appearance:none}
.submit-btn:active{opacity:.85}
.submit-btn:disabled{opacity:.6}
.result{text-align:center;padding:16px;margin:16px;border-radius:12px;display:none;font-size:15px;line-height:1.6}
.result.ok{background:#e8f5e9;color:#2e7d32;display:block}
.result.err{background:#fce4ec;color:#c62828;display:block}
.result.loading{background:#e3f2fd;color:#1565c0;display:block}
.wecom-modal{position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.6);display:none;align-items:center;justify-content:center;z-index:1000;padding:20px}
.wecom-modal.active{display:flex}
.wecom-modal .box{background:#fff;border-radius:16px;padding:24px;text-align:center;max-width:340px;width:100%}
.wecom-modal .box h3{color:#1a73e8;margin-bottom:12px;font-size:18px}
.wecom-modal .box img{width:220px;height:auto;border-radius:8px;margin:8px 0}
.wecom-modal .box .tip{color:#666;font-size:13px;margin-bottom:16px}
.wecom-modal .box .btn{display:inline-block;background:linear-gradient(135deg,#1a73e8,#0d47a1);color:#fff;padding:12px 32px;border-radius:10px;text-decoration:none;font-size:15px;font-weight:700;cursor:pointer;border:none}
</style>
</head>
<body>

<div class="top">
  <div class="label">推荐人：__NAME__</div>
  <h2>EUGOOO 入驻申请表</h2>
  <div class="ref">推荐编码已自动锁定 <strong>__ID__</strong></div>
</div>

__SCRIPT_CARD__

<div class="card">
  <h3>填写入驻信息</h3>
  <form id="regForm">
    <div class="field">
      <label>推荐人编码（自动锁定）</label>
      <input type="text" value="__ID__" readonly class="locked" id="refCode">
    </div>
    <div class="field">
      <label><span class="req">*</span> 姓名</label>
      <input type="text" id="fName" required placeholder="请输入姓名">
    </div>
    <div class="field">
      <label><span class="req">*</span> 手机号</label>
      <input type="tel" id="fPhone" required placeholder="请输入手机号">
    </div>
    <div class="field">
      <label>微信号</label>
      <input type="text" id="fWechat" placeholder="请输入微信号">
    </div>
    <div class="field">
      <label>所在城市</label>
      <input type="text" id="fCity" placeholder="请输入所在城市">
    </div>
    <div class="field">
      <label>所在学校</label>
      <input type="text" id="fSchool" placeholder="请输入学校名称">
    </div>
    <div class="field">
      <label>是否已了解 eugooo.com 平台</label>
      <select id="fKnown">
        <option value="">请选择</option>
        <option value="是">是</option>
        <option value="否">否</option>
      </select>
    </div>
    <div class="field">
      <label>用户类型</label>
      <select id="fUserType">
        <option value="">请选择</option>
        <option value="微信用户">微信用户</option>
        <option value="大学生">大学生</option>
        <option value="创业者">创业者</option>
        <option value="其他">其他</option>
      </select>
    </div>
    <button type="button" class="submit-btn" id="submitBtn" onclick="doSubmit()">提交入驻申请</button>
  </form>
</div>

<div class="result" id="result"></div>

<div class="wecom-modal" id="wecomModal">
  <div class="box">
    <h3>🎉 提交成功！</h3>
    <div class="tip">请长按二维码识别，添加 <strong>__NAME__</strong> 的企业微信</div>
    __WECOM_IMG__
    <button class="btn" onclick="closeWecomModal()">我知道了</button>
  </div>
</div>

<script>
var SUBMITTING = false;
var SUBMISSIONS_URL = '__API_BASE__';

function showResult(cls, msg){
  var r = document.getElementById('result');
  r.className = 'result ' + cls;
  r.textContent = msg;
}

function openWecomModal(){
  document.getElementById('wecomModal').classList.add('active');
}

function closeWecomModal(){
  document.getElementById('wecomModal').classList.remove('active');
}

function doSubmit(){
  if(SUBMITTING) return;
  var n = document.getElementById('fName').value.trim();
  var p = document.getElementById('fPhone').value.trim();
  if(!n || !p){ alert('\\u8bf7\\u81f3\\u5c11\\u586b\\u5199\\u59d3\\u540d\\u548c\\u624b\\u673a\\u53f7'); return; }

  SUBMITTING = true;
  var btn = document.getElementById('submitBtn');
  btn.disabled = true;
  btn.textContent = '\\u63d0\\u4ea4\\u4e2d...';
  showResult('loading', '\\u6b63\\u5728\\u63d0\\u4ea4\\uff0c\\u8bf7\\u7a0d\\u5019...');

  var entry = {
    refCode: document.getElementById('refCode').value,
    name: n,
    phone: p,
    wechat: document.getElementById('fWechat').value.trim(),
    city: document.getElementById('fCity').value.trim(),
    school: document.getElementById('fSchool').value.trim(),
    known: document.getElementById('fKnown').value,
    userType: document.getElementById('fUserType').value,
    time: new Date().toLocaleString('zh-CN', {timeZone:'Asia/Shanghai'})
  };

  fetch(SUBMISSIONS_URL)
    .then(function(r){ return r.json(); })
    .then(function(data){
      if(!Array.isArray(data)) data = [];
      data.push(entry);
      return fetch(SUBMISSIONS_URL, {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify(data)
      });
    })
    .then(function(r){ return r.json(); })
    .then(function(){
      showResult('ok', '\\u63d0\\u4ea4\\u6210\\u529f\\uff01\\n\\u63a8\\u8350\\u4eba\\u7f16\\u7801 __ID__ \\u5df2\\u81ea\\u52a8\\u8bb0\\u5f55\\u3002\\n\\u6b22\\u8fce\\u52a0\\u5165 EUGOOO\\uff01');
      document.getElementById('regForm').style.display = 'none';
      SUBMITTING = false;
      btn.disabled = false;
      btn.textContent = '\\u63d0\\u4ea4\\u5165\\u9a7b\\u7533\\u8bf7';
      openWecomModal();
    })
    .catch(function(){
      SUBMITTING = false;
      btn.disabled = false;
      btn.textContent = '\\u63d0\\u4ea4\\u5165\\u9a7b\\u7533\\u8bf7';
      showResult('err', '\\u63d0\\u4ea4\\u5931\\u8d25\\uff0c\\u8bf7\\u68c0\\u67e5\\u7f51\\u7edc\\u540e\\u91cd\\u8bd5');
    });
}
</script>
</body>
</html>'''

ADMIN_TPL = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>EUGOOO · 数据后台</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:"Microsoft YaHei","PingFang SC",sans-serif;background:#f5f7fa;min-height:100vh;color:#222}
header{background:linear-gradient(135deg,#1a73e8,#0d47a1);color:#fff;padding:20px;text-align:center}
header h1{font-size:20px;margin-bottom:4px}
header p{font-size:13px;opacity:.85}
.stats{display:flex;gap:12px;padding:16px;flex-wrap:wrap}
.stat{flex:1;min-width:80px;background:#fff;border-radius:12px;padding:16px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.06)}
.stat b{font-size:28px;color:#1a73e8;display:block}
.stat span{font-size:12px;color:#888;margin-top:4px;display:block}
.controls{padding:0 16px 12px;display:flex;gap:10px;flex-wrap:wrap}
.controls button{padding:10px 20px;border:none;border-radius:10px;font-size:14px;cursor:pointer;font-weight:600}
.btn-refresh{background:#1a73e8;color:#fff}
.btn-csv{background:#4caf50;color:#fff}
.btn-clear{background:#e53935;color:#fff}
.table-wrap{overflow-x:auto;padding:0 16px 16px}
table{width:100%;border-collapse:collapse;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.06);font-size:13px}
th,td{padding:10px 8px;text-align:left;border-bottom:1px solid #eee;white-space:nowrap}
th{background:#f0f4ff;color:#1a73e8;font-weight:600;position:sticky;top:0}
tr:last-child td{border-bottom:none}
.empty{padding:40px;text-align:center;color:#999;font-size:15px}
</style>
</head>
<body>
<header>
  <h1>EUGOOO 入驻数据后台</h1>
  <p>实时数据 · 自动刷新 · 支持导出CSV</p>
</header>
<div class="stats" id="stats"></div>
<div class="controls">
  <button class="btn-refresh" onclick="load()">刷新数据</button>
  <button class="btn-csv" onclick="downloadCSV()">导出 CSV</button>
  <button class="btn-clear" onclick="clearData()">清空数据</button>
</div>
<div class="table-wrap" id="tableArea"></div>

<script>
var SUBMISSIONS_URL = '__API_BASE__';
var DATA = [];

function load(){
  document.getElementById('tableArea').innerHTML = '<div class="empty">\\u52a0\\u8f7d\\u4e2d...</div>';
  fetch(SUBMISSIONS_URL)
    .then(function(r){ return r.json(); })
    .then(function(d){
      DATA = Array.isArray(d) ? d : [];
      render();
    })
    .catch(function(){
      document.getElementById('tableArea').innerHTML = '<div class="empty">\\u52a0\\u8f7d\\u5931\\u8d25\\uff0c\\u8bf7\\u68c0\\u67e5\\u7f51\\u7edc\\u6216\\u5237\\u65b0\\u91cd\\u8bd5<br><small>API: ' + SUBMISSIONS_URL + '</small></div>';
    });
}

function render(){
  var uniqueRef = new Set(DATA.map(function(r){ return r.refCode; })).size;
  var lastTime = DATA.length ? DATA[DATA.length-1].time : '-';
  document.getElementById('stats').innerHTML =
    '<div class="stat"><b>' + DATA.length + '</b><br><span>\\u603b\\u63d0\\u4ea4\\u6570</span></div>' +
    '<div class="stat"><b>' + uniqueRef + '</b><br><span>\\u63a8\\u8350\\u4eba</span></div>' +
    '<div class="stat"><b>' + lastTime + '</b><br><span>\\u6700\\u540e\\u66f4\\u65b0</span></div>';

  if(!DATA.length){
    document.getElementById('tableArea').innerHTML = '<div class="empty">\\u6682\\u65e0\\u63d0\\u4ea4\\u6570\\u636e</div>';
    return;
  }
  var rows = DATA.map(function(r){
    return '<tr><td>'+(r.refCode||'')+'</td><td>'+(r.name||'')+'</td><td>'+(r.phone||'')+'</td><td>'+(r.wechat||'')+'</td><td>'+(r.city||'')+'</td><td>'+(r.school||'')+'</td><td>'+(r.known||'')+'</td><td>'+(r.userType||'')+'</td><td>'+(r.time||'')+'</td></tr>';
  }).join('');
  document.getElementById('tableArea').innerHTML =
    '<table><thead><tr><th>\\u63a8\\u8350\\u4eba\\u7f16\\u7801</th><th>\\u59d3\\u540d</th><th>\\u624b\\u673a\\u53f7</th><th>\\u5fae\\u4fe1\\u53f7</th><th>\\u57ce\\u5e02</th><th>\\u5b66\\u6821</th><th>\\u662f\\u5426\\u4e86\\u89e3</th><th>\\u7528\\u6237\\u7c7b\\u578b</th><th>\\u65f6\\u95f4</th></tr></thead><tbody>'+rows+'</tbody></table>';
}

function downloadCSV(){
  var csv = '\\uFEFF\\u63a8\\u8350\\u4eba\\u7f16\\u7801,\\u59d3\\u540d,\\u624b\\u673a\\u53f7,\\u5fae\\u4fe1\\u53f7,\\u57ce\\u5e02,\\u5b66\\u6821,\\u662f\\u5426\\u4e86\\u89e3,\\u7528\\u6237\\u7c7b\\u578b,\\u65f6\\u95f4\\n';
  DATA.forEach(function(r){
    csv += [r.refCode||'', r.name||'', r.phone||'', r.wechat||'', r.city||'', r.school||'', r.known||'', r.userType||'', r.time||''].join(',') + '\\n';
  });
  var blob = new Blob([csv], {type:'text/csv;charset=utf-8'});
  var a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = 'eugooo_submissions.csv';
  a.click();
}

function clearData(){
  if(!confirm('\\u786e\\u5b9a\\u8981\\u6e05\\u7a7a\\u6240\\u6709\\u63d0\\u4ea4\\u6570\\u636e\\uff1f\\u6b64\\u64cd\\u4f5c\\u4e0d\\u53ef\\u64a4\\u9500\\u3002')) return;
  fetch(SUBMISSIONS_URL, {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: '[]'
  })
    .then(function(r){ return r.json(); })
    .then(function(d){
      alert(d.ok ? '\\u5df2\\u6e05\\u7a7a' : '\\u6e05\\u7a7a\\u5931\\u8d25');
      load();
    })
    .catch(function(){ alert('\\u6e05\\u7a7a\\u5931\\u8d25\\uff0c\\u8bf7\\u68c0\\u67e5\\u7f51\\u7edc'); });
}

load();
setInterval(load, 30000);
</script>
</body>
</html>'''

INDEX_TPL = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>EUGOOO · 专属分享码汇总</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:"Microsoft YaHei","PingFang SC",sans-serif;background:#f5f7fa;min-height:100vh}
.header{background:linear-gradient(135deg,#1a73e8,#0d47a1);color:#fff;text-align:center;padding:28px 20px}
.header h1{font-size:22px;margin-bottom:6px}
.header p{font-size:13px;opacity:.85}
.grid{display:flex;flex-wrap:wrap;gap:16px;padding:20px;justify-content:center}
.card{background:#fff;width:200px;padding:16px;border-radius:16px;text-align:center;box-shadow:0 2px 12px rgba(0,0,0,.06)}
.card .name{font-size:16px;font-weight:700;margin:8px 0;color:#222}
.card .code{font-size:12px;color:#888;margin-bottom:8px}
.card img{width:150px;height:150px}
.card .link{font-size:11px;color:#1a73e8;word-break:break-all;margin-top:8px}
.edit-area{margin:16px;padding:12px;background:#fff3e0;border-radius:12px;text-align:center;font-size:13px}
.edit-area input{padding:8px 12px;border:1px solid #ddd;border-radius:8px;font-size:13px;width:280px}
.edit-area button{padding:8px 16px;border:none;border-radius:8px;background:#1a73e8;color:#fff;font-size:13px;cursor:pointer;margin-left:8px}
.tip{text-align:center;color:#999;font-size:12px;padding:16px}
</style>
</head>
<body>
<div class="header">
  <h1>EUGOOO 跨境分销平台 · 专属分享码 V7</h1>
  <p>共 __COUNT__ 人 · 扫码打开专属表单 · 推荐编码自动锁定 · 数据实时入库 · __TIME__</p>
</div>
<div class="edit-area">
  <span>API 地址：</span>
  <input id="apiInput" value="__API_BASE__">
  <button onclick="updateAPI()">更新</button>
  <span style="color:#888;margin-left:12px;font-size:12px">修改后可重新生成所有页面</span>
</div>
<div class="grid">
__CARDS__
</div>
<div class="tip">每人独立二维码，扫码直达专属入驻表单 · 推荐人编码自动锁定不可修改 · 数据实时存入后台 · 管理员查看 <a href="__ADMIN_URL__" target="_blank">数据后台</a></div>
</body>
</html>'''

# ===== Generate individual form pages =====
for p in people:
    script_html = p['script'].replace('\n', '<br>') if p['script'] else '暂无话术'
    card_file = available_cards.get(p['name'])

    if p['script'] or card_file:
        if card_file:
            script_card_html = '<div class="card card-script">'
            script_card_html += f'<div class="wecom-col"><img src="{card_file}" alt="{p["name"]}企业微信"><div class="cname">{p["name"]}</div><div class="cdesc">扫码添加企业微信</div></div>'
            script_card_html += f'<div class="script-col"><h3>推广话术</h3><div class="msg">{script_html}</div></div>'
            script_card_html += '</div>'
        else:
            script_card_html = f'<div class="card"><h3>推广话术</h3><div class="msg">{script_html}</div></div>'
    else:
        script_card_html = ''

    # 企业微信名片弹窗图片
    card_file = available_cards.get(p['name'])
    if card_file:
        wecom_img_html = f'<img src="{card_file}" alt="{p["name"]}企业微信">'
    else:
        wecom_img_html = '<div style="color:#999;font-size:13px;padding:20px;">暂无企业微信名片二维码</div>'

    html = (FORM_TPL
            .replace('__NAME__', p['name'])
            .replace('__ID__', p['id'])
            .replace('__API_BASE__', API_BASE)
            .replace('__SCRIPT_CARD__', script_card_html)
            .replace('__WECOM_IMG__', wecom_img_html))

    fname = p['id'] + '.html'
    with open(os.path.join(OUT, fname), 'w', encoding='utf-8') as f:
        f.write(html)
    card_tag = ' [with card]' if card_file else ''
    print(f'  Wrote {fname}{card_tag}')

# ===== Generate admin page =====
admin_html = ADMIN_TPL.replace('__API_BASE__', API_BASE)
with open(os.path.join(OUT, 'admin.html'), 'w', encoding='utf-8') as f:
    f.write(admin_html)
print('  Wrote admin.html')

# ===== Generate QR codes =====
for p in people:
    url = f'{BASE_URL}/{p["id"]}.html'
    qr = qrcode.QRCode(version=2, box_size=10, border=2)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white').convert('RGB')

    from PIL import Image, ImageDraw, ImageFont
    label_h = 40
    new_img = Image.new('RGB', (img.width, img.height + label_h), 'white')
    new_img.paste(img, (0, 0))
    draw = ImageDraw.Draw(new_img)
    label = f'{p["name"]} ({p["id"]})'
    try:
        font = ImageFont.truetype('msyh.ttc', 16)
    except:
        font = ImageFont.load_default()
    bbox = draw.textbbox((0, 0), label, font=font)
    tw = bbox[2] - bbox[0]
    draw.text(((img.width - tw) // 2, img.height + 10), label, fill='black', font=font)

    fname = f'{p["id"]}_{p["name"]}.png'
    new_img.save(os.path.join(QR_DIR, fname))
    print(f'  QR: {fname} -> {url}')

# ===== Generate index page =====
cards = ''
for p in people:
    cards += f'''<div class="card">
  <img src="qrcodes/{p['id']}_{p['name']}.png" alt="{p['name']}">
  <div class="name">{p['name']}</div>
  <div class="code">编码 {p['id']}</div>
  <div class="link">{p['id']}.html</div>
</div>\n'''

from datetime import datetime
ts = datetime.now().strftime('%Y-%m-%d %H:%M')
index = (INDEX_TPL
         .replace('__COUNT__', str(len(people)))
         .replace('__TIME__', ts)
         .replace('__API_BASE__', API_BASE)
         .replace('__CARDS__', cards)
         .replace('__ADMIN_URL__', f'{BASE_URL}/admin.html'))

with open(os.path.join(OUT, 'index.html'), 'w', encoding='utf-8') as f:
    f.write(index)
print(f'  Wrote index.html ({len(people)} people)')

print(f'\nDone! API: {API_BASE}')
